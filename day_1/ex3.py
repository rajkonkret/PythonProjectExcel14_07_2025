# xlwt, xlrd
import openpyxl
from pprint import pprint

wb = openpyxl.load_workbook("../data/videogamesales.xlsx")
ws = wb.active
print(ws)  # <Worksheet "vgsales">

ws = wb['vgsales']

print("Total number of row: ", ws.max_row)
print("Total number of columns: ", ws.max_column)
# Total number of row:  16328
# Total number of columns:  10

print("Value in cell A1 is:", ws['A1'].value)  # Value in cell A1 is: Rank

values = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column)]
print(values)
# ['Rank', 'Name', 'Platform', 'Year', 'Genre', 'Publisher', 'NA_Sales', 'EU_Sales', 'JP_Sales']

data = [ws.cell(row=i, column=2).value for i in range(2, 12)]
print(data)
# ['Wii Sports', 'Super Mario Bros.', 'Mario Kart Wii', 'Wii Sports Resort', 'Pokemon Red/Pokemon Blue', 'Tetris',
#  'New Super Mario Bros.', 'Wii Play', 'New Super Mario Bros. Wii', 'Duck Hunt']

pprint(data)

my_list = list()  # pusta lista

for value in ws.iter_rows(
        min_row=1, max_row=11, min_col=1, max_col=6,
        values_only=True):
    my_list.append(value)

pprint(my_list)

for ele1, ele2, ele3, ele4, ele5, ele6 in my_list:
    print(f"{ele1:<8}{ele2:>35}{ele3:^10}{ele4:<10}{ele5:<15}{ele6:<15}")

ws['K1'] = "Sum of Sales"

wb.save("videogamesales2.xlsx")
wb.close()
