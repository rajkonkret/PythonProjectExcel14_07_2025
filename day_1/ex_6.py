import openpyxl

wb = openpyxl.load_workbook('videogamesales2.xlsx')
ws = wb.active

ws = wb['vgsales']

# new_row = (1, 'The Legend of Zelda', "Wii", 1986, "Action", "Nintendo", 3.74, 0.93, 1.69, 0.14, 6.51)
# ws.append(new_row)
#
# wb.save("videogamesales2.xlsx")
# wb.close()

# odczyt
values = [ws.cell(row=ws.max_row, column=i).value for i in range(1, ws.max_column + 1)]
print(values)
# [1, 'The Legend of Zelda', 'Wii', 1986, 'Action', 'Nintendo', 3.74, 0.93, 1.69, 0.14, 6.51]

# usunięcie jednego wiersza (ostatniego)
ws.delete_rows(ws.max_row, 1)

wb.save('videogamesales2.xlsx')
wb.close()
