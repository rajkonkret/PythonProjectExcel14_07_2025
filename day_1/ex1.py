from openpyxl import Workbook, load_workbook

#  pip install openpyxl

wb = Workbook()
ws = wb.active

ws['A1'] = 42
ws.append([1, 2, 3])  # dodaj

wb.save('sample.xlsx')
wb.close()
print("Excel file has been created succesfully")

workbook = load_workbook('sample.xlsx')
sheet = workbook.active
print(sheet)
# <Worksheet "Sheet">

print(sheet['A1'].value)  # 42

# indent rainbow
for row in sheet.iter_rows(min_row=1, max_row=3):
    for cell in row:
        print(cell.value)
# 42
# None
# None
# 1
# 2
# 3
# None
# None
# None
