import openpyxl

wb = openpyxl.load_workbook('videogamesales2.xlsx')

ws = wb['vgsales']

ws['P1'] = 'Average Sales'
ws['P2'] = '=AVERAGE(K2:K16599)'

wb.save('videogamesales2.xlsx')
wb.close()

ws['Q1'] = "Number of populated cells"
ws['Q2'] = "=COUNTA(E2:E16599)"

wb.save('videogamesales2.xlsx')
wb.close()

ws['R1'] = "Number of Rows with Sports Genre"
ws['R2'] = '=COUNTIF(E2:E16599, "Sports")'

wb.save('videogamesales2.xlsx')
wb.close()

ws['S1'] = "Total Sports Sales"
ws['S2'] = '=SUMIF(E2:E16599, "Sports", K2:K16599)'

wb.save('videogamesales2.xlsx')
wb.close()

ws['T1'] = "Rounded Sum of Sports Sales"
ws['T2'] = '=CEILING(S2,25)'  # skok co 25 np: 1308.9 -> 1325

wb.save('videogamesales2.xlsx')
wb.close()
