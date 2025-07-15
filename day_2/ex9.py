import openpyxl

from openpyxl.styles import Font, colors, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.load_workbook('videogamesales2.xlsx')
ws = wb.active

# print(ws.title)
# ws.title = "Video Games Sales Data"
# wb.save('videogamesales2.xlsx')
# wb.close()
# ctrl / - komentarz

sheet_names = wb.sheetnames
print(sheet_names)
# ['Video Games Sales Data', 'Total Sales by Genre', 'Breakdown of Sales by Genre', 'Breakdown of Sales by Year']
sheet = wb[sheet_names[1]]
print(sheet)  # <Worksheet "Total Sales by Genre">
wb.active = wb.index(sheet)

ws = wb.active
print(ws.title)  # Total Sales by Genre

ws = wb['Video Games Sales Data']
# ws['A1'].font = Font(bold=True, size=12)
#
# for cell in ws[1:1]:
#     cell.font = Font(bold=True, size=12)
#
# wb.save("videogamesales2.xlsx")
# wb.close()

print(wb.sheetnames)
# ['Video Games Sales Data', 'Total Sales by Genre', 'Breakdown of Sales by Genre', 'Breakdown of Sales by Year']

# utworzenie nowego arkusza
# wb.create_sheet('Empty Sheet')
# print(wb.sheetnames)
# # ['Video Games Sales Data', 'Total Sales by Genre', 'Breakdown of Sales by Genre', 'Breakdown of Sales by Year', 'Empty Sheet']
#
# wb.save('videogamesales2.xlsx')
# wb.close()

# wb.remove(wb['Empty Sheet'])
# print(wb.sheetnames)
# # ['Video Games Sales Data', 'Total Sales by Genre', 'Breakdown of Sales by Genre', 'Breakdown of Sales by Year']
# wb.save('videogamesales2.xlsx')
# wb.close()

wb.copy_worksheet(wb["Video Games Sales Data"])
wb.save('vgsales.xlsx')
wb.close()